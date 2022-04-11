import os
from statistics import median_grouped
from tkinter import messagebox
from threading import Thread
import openpyxl  
from openpyxl.styles import PatternFill,Border, Side, Font, fills, Alignment, Protection, borders
import datetime
import time
import subprocess
from database import saveBillLocation, saveCustomerData, saveGstData
from baseIntialization import UiFields
from backend import convert
from printer import printBill


def generateBill(u : UiFields):
    if u.bill_generated:
        if u.mobile_txt.get()!='':
            print_bill(u)
        return
    u.bill_generated = True
    wb = openpyxl.Workbook()
    sh1 = wb.active
    sh1.title = 'bill'
    daten = datetime.datetime.now()
# ====================================================shop details==================================================================
    sh1['D1'] = "    GIRIDHARI JEWELLERY" 
    sh1['D1'].font = Font(name='Copperplate Gothic Bold',size=35,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='c24949')

    sh1['D2'] = "    Badheibanka Chawk, Old Town, Bhubaneswar, Mob.: 9090280083,9861230757"
    sh1['D2'].font = Font(name='times new rommon',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='009966')#'064635')

    sh1['E3'] = "            916 GOLD AND SILVER ORNAMENT"
    sh1['E3'].font = Font(name='times new rommon',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='c85c5c')
    
    sh1['F4'] = "                 Tax Invoice"
    sh1['F4'].font = Font(name='times new rommon',size=16,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='c85c5c')

    hallmark_logo = openpyxl.drawing.image.Image(r"hallmark.png")      
    hallmark_logo.height = 110  
    hallmark_logo.width = 105
    sh1.add_image(hallmark_logo, "K1")



    shop_logo = openpyxl.drawing.image.Image(r"shopLogo.png")      
    shop_logo.height = 110  
    shop_logo.width = 105
    sh1.add_image(shop_logo, "A1")

# ======================================================customer details==================================================
    
    sh1['A5'] = 'GSTIN - 21AYRPK4931F1ZH'
    sh1['A5'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)

    sh1['F5'] = "Bill. No.: "
    sh1['F5'].font = Font(name='times new rommon',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['G5'] = u.bill_txt
    sh1['G5'].font = Font(name='times new rommon',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)

    sh1['I5'] = 'Date: '+ daten.strftime("%d-%b-%y - (%A)")
    sh1['I5'].font = Font(name='times new rommon',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)

    sh1['A6'] = "Name: "
    sh1['A6'].font = Font(name='times new rommon',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['B6'] = u.name_txt.get()
    sh1['B6'].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)

    sh1['E6'] = 'Mobile No.'
    sh1['E6'].font = Font(name='times new rommon',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['F6'] = '  '+ u.mobile_txt.get()
    sh1['F6'].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)

    sh1['H6'] = 'Addhar No.'
    sh1['H6'].font = Font(name='times new rommon',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['I6'] = u.addhar_txt.get()
    sh1['I6'].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)

    sh1['F7'] = "Gold Rate: "
    sh1['F7'].font = Font(name='times new rommon',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['G7'] = u.gold_rate*10
    sh1['G7'].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)
    
    
# ===========================================product details======================================================================
    
    sh1['A8'] = 'Si.'
    sh1['A8'].font = Font(name='times new romman',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['B8'] = 'Description Of Goods'
    sh1['B8'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['E8'] = 'HSN'
    sh1['E8'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['E9'] = 'SAC'
    sh1['E9'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    
    sh1['F8'] = 'Weight'
    sh1['F8'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)

    sh1['G8'] = 'MC'
    sh1['G8'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)

    sh1['H8'] = 'Unit'
    sh1['H8'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['H9'] = 'Price'
    sh1['H9'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)

    sh1['I8'] = 'CGST'
    sh1['I8'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['I9'] = '1.5%'
    sh1['I9'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)

    sh1['J8'] = 'SGST'
    sh1['J8'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['J9'] = '1.5%'
    sh1['J9'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)

    sh1['K8'] = 'Net'
    sh1['K8'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
    sh1['K9'] = 'Total'
    sh1['K9'].font = Font(name='times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)

    sh1['E10'] = '7113'
    sh1['E10'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

    for i in range(1,10):
        if(u.des_txt[i-1].get()!= ""):
            sh1['A'+str(i+9)] = i
            sh1['A'+str(i+9)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

            sh1['B'+str(i+9)] = u.des_txt[i-1].get()
            sh1['B'+str(i+9)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

            sh1['F'+str(i+9)] = u.wt_txt[i-1].get() + 'gm'
            sh1['F'+str(i+9)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

            sh1['G'+str(i+9)] = u.mc_txt[i-1].get()
            sh1['G'+str(i+9)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

            sh1['H'+str(i+9)] = u.unit_txt[i-1].get()
            sh1['H'+str(i+9)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

            sh1['I'+str(i+9)] = u.cgst_txt[i-1].get()
            sh1['I'+str(i+9)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

            sh1['J'+str(i+9)] = u.sgst_txt[i-1].get()
            sh1['J'+str(i+9)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

            sh1['K'+str(i+9)] = u.net_txt[i-1].get()
            sh1['K'+str(i+9)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
            if(u.wt_txt[i-1].get()!='' and u.net_txt[i-1].get()!=''):
                saveGstData(float(u.wt_txt[i-1].get()),u.des_txt[i-1].get(),u.gold_rate,u.total_taxable_amt[i-1],float(u.cgst_txt[i-1].get()),float(u.sgst_txt[i-1].get()),float(u.net_txt[i-1].get()))

# ===========================================================old gold============================================================
        sh1['A19'] = 'Si.'
        sh1['A19'].font = Font(name='Times new romman',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

        
        sh1['D19'] = "Old Jewellery"
        sh1['D19'].font = Font(name='Times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)

        sh1['I19'] = "Weight"
        sh1['I19'].font = Font(name='Times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
        
        sh1['J19'] = "Unit Price"
        sh1['J19'].font = Font(name='Times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)

        sh1['K19'] = "Total"
        sh1['K19'].font = Font(name='Times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
        for i in range(1,4):
            if u.oldtotal_txt[i-1].get()!="":
                sh1['A2'+str(i-1)] = i
                sh1['A2'+str(i-1)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
                
                sh1['B2'+str(i-1)] = u.oldDesc_txt[i-1].get()
                sh1['B2'+str(i-1)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
                
                sh1['I2'+str(i-1)] = u.oldwe_txt[i-1].get()
                sh1['I2'+str(i-1)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
                
                sh1['J2'+str(i-1)] = u.oldunit_txt[i-1].get()
                sh1['J2'+str(i-1)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
                
                sh1['K2'+str(i-1)] = u.oldtotal_txt[i-1].get()
                sh1['K2'+str(i-1)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
                
#====================================================ADDITION OR DEDUCTION================================================================================================================
        sh1['A24'] = 'Si.'
        sh1['A24'].font = Font(name='Times new romman',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
        
        sh1['C24'] = "Other Addition Or Deduction"
        sh1['C24'].font = Font(name='Times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
        
        sh1['K24'] = "Amount"
        sh1['K24'].font = Font(name='Times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.green)
        
        for i in range(1,4):
            if u.addtotal_txt[i-1].get()!="":
                sh1['A2'+str(i+4)] = i
                sh1['A2'+str(i+4)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
                
                sh1['B2'+str(i+4)] = u.addDesc_txt[i-1].get()
                sh1['B2'+str(i+4)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
                
                sh1['K2'+str(i+4)] = u.addtotal_txt[i-1].get()
                sh1['K2'+str(i+4)].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
                
        
            

    # ===================================================mode of payment============================================


        sh1['E28'] = "Mode Of Payment"
        sh1['E28'].font = Font(name='Times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)
        mode = u.mode
        if(u.charge.get()!=''):
            mode = mode + ':    '+ u.charge.get()
        sh1['I28'] = mode
        sh1['I28'].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

# =====================================================Total paid=======================================================

        sh1['A29'] = "Total in Words"
        sh1['A29'].font = Font(name='Times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)
        try:
            tot = int(float(u.total.get()))
            sh1['C29'] = convert(tot)
            sh1['C29'].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
        except Exception as e:
            print("There is a error in convert : {0}".format(e))

        sh1['J29'] = "Total Paid : "
        sh1['J29'].font = Font(name='Times new romman',size=16,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)
        sh1['K29'] = u.total.get()
        sh1['K29'].font = Font(name='arial',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)
# =================================================Terms and conditions===================================================
        # sh1['B30']="Terms and conditions"
        # sh1['B30'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)

        # sh1['A31'] = " 1. Before taking delivery Customer should check the ornaments by piece and weight. "
        # sh1['A31'].font = Font(name='Times new romman',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

        # sh1['A32'] = " 2. Exchange your ornaments within 7 days in good condition."
        # sh1['A32'].font = Font(name='Times new romman',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

        # sh1['A33'] = " 3. We are not responsible for any damage or brakage after delivery"
        # sh1['A33'].font = Font(name='Times new romman',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)

        # sh1['A34'] = " 4. Any type of repairing of our ornament is FREE"
        # sh1['A34'].font = Font(name='Times new romman',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.blue)


        # sh1['L33'] = "Signature"
        # sh1['L33'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)





    thin = Side(border_style='thin',color='808080')#='000066')
    medium = Side(border_style='medium',color='808080')#'000066')
# ========================================================creating borders====================================================         

    fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00',end_color='00FFFF00')



   
    # ================================================== 5 =====================================================
    for i in range(2,11):
        sh1.cell(row=5, column=i).border=Border(top=medium,bottom=thin)
    sh1.cell(row=5,column=1).border=Border(left=medium,top=medium,bottom=thin)
    sh1.cell(row=5,column=11).border=Border(right=medium,top=medium,bottom=thin)
    
    # ==================================================== 6 =====================================================

    sh1.cell(row = 6,column = 1).border=Border(left=medium,top=thin)
    sh1.cell(row = 6,column = 11).border=Border(right=medium,top=thin)

    # ======================================================= 7 ====================================================

    sh1.cell(row = 7,column = 11).border=Border(right=medium,bottom=thin)
    sh1.cell(row = 7,column = 1).border=Border(left=medium,bottom=thin)


    # ======================================================= 8 =====================================================
    sh1.cell(row = 8,column = 1).border=Border(left=medium,top=thin)
    sh1.cell(row = 8,column = 2).border=Border(left=thin,top=thin)
    sh1.cell(row = 8,column = 3).border=Border(top=thin)
    sh1.cell(row = 8,column = 4).border=Border(right=thin,top=thin)
    sh1.cell(row = 8,column = 5).border=Border(right=thin,top=thin)

    sh1.cell(row = 8,column = 6).border=Border(left=thin,top=thin)
    sh1.cell(row = 8,column = 7).border=Border(left=thin,top=thin)
    sh1.cell(row = 8,column = 8).border=Border(left=thin,top=thin)
    sh1.cell(row = 8,column = 9).border=Border(left=thin,top=thin)
    sh1.cell(row = 8,column = 10).border=Border(right=thin,top=thin)

    sh1.cell(row = 8,column = 11).border=Border(left=thin,top=thin,right=medium)
    # sh1.cell(row = 8,column = 12).border=Border(right=thin,top=thin)
    # sh1.cell(row = 8,column = 13).border=Border(top=thin)
    sh1.cell(row = 8,column = 11).border=Border(right=medium,top=thin)
    
    # ===================================================== 9 ======================================================
    
    sh1.cell(row = 9,column = 1).border=Border(left=medium,bottom=thin)
    sh1.cell(row = 9,column = 2).border=Border(left=thin,bottom=thin)
    sh1.cell(row = 9,column = 3).border=Border(bottom=thin)
    sh1.cell(row = 9,column = 4).border=Border(bottom=thin,right=thin)
    sh1.cell(row = 9,column = 5).border=Border(bottom=thin,right=thin)
    sh1.cell(row = 9,column = 6).border=Border(left=thin,bottom=thin)
    sh1.cell(row = 9,column = 7).border=Border(left=thin,bottom=thin)
    sh1.cell(row = 9,column = 8).border=Border(left=thin,bottom=thin)
    sh1.cell(row = 9,column = 9).border=Border(left=thin,bottom=thin)
    sh1.cell(row = 9,column = 10).border=Border(right=thin,bottom=thin)
    sh1.cell(row = 9,column = 11).border=Border(left=thin,bottom=thin,right=medium)
    # sh1.cell(row = 9,column = 12).border=Border(right=thin,bottom=thin)
    # sh1.cell(row = 9,column = 13).border=Border(bottom=thin)
    sh1.cell(row = 9,column = 11).border=Border(right=medium,bottom=thin)


    # ========================================================== 10 =======================================================

    sh1.cell(row = 10,column = 1).border = Border(left=medium)
    for i in range(2,12):
        if(i!=3 and i!=4 and i!=5 ):
            sh1.cell(row = 10,column = i).border = Border(left=thin)
    sh1.cell(row = 10,column = 4).border=Border(right=thin,top=thin)

    # ============================================================= 10 to 18 ===================================================
    
    # for i in range(10,11):
    sh1.cell(row = 10,column = 11).border=Border(left=thin,right=medium)

    for i in range(11,18):
        sh1.cell(row = i,column = 1).border = Border(left=medium)
        sh1.cell(row = i,column = 4).border = Border(right=thin)
        sh1.cell(row = i,column = 11).border = Border(left=thin,right=medium)

    for j in range(1,12):
        if(j==2 or j==6 or j==7 or j==8 or j==9 ):  
            for i in range(11,18):
                sh1.cell(row = i,column = j).border = Border(left=thin)
    for i in range(11,18):
        sh1.cell(row = i,column = 10).border = Border(left=thin,right=thin)

    sh1.cell(row = 18,column = 1).border=Border(left=medium,bottom=thin,right=thin)

    sh1.cell(row = 18,column = 2).border=Border(left=thin,bottom=thin)
    sh1.cell(row = 18,column = 3).border=Border(bottom=thin)
    sh1.cell(row = 18,column = 4).border=Border(bottom=thin,right=thin)
    sh1.cell(row = 18,column = 5).border=Border(bottom=thin,right=thin)
    sh1.cell(row = 18,column = 6).border=Border(left=thin,bottom=thin,right=thin)
    sh1.cell(row = 18,column = 7).border=Border(left=thin,bottom=thin,right=thin)
    sh1.cell(row = 18,column = 8).border=Border(left=thin,bottom=thin,right=thin)
    sh1.cell(row = 18,column = 9).border=Border(left=thin,bottom=thin)
    sh1.cell(row = 18,column = 10).border=Border(bottom=thin,right=thin,left=thin)
    sh1.cell(row = 18,column = 11).border=Border(left=thin,bottom=thin,right=medium)
    # sh1.cell(row = 18,column = 12).border=Border(bottom=thin,right=thin)
    # sh1.cell(row = 18,column = 13).border=Border(left=thin,bottom=thin)    
    # sh1.cell(row = 18,column = 15).border=Border(right=medium,bottom=thin)    
    

    # ======================================================= 19 ==========================================================

    # sh1.cell(row = 19,column = 14).border=Border(right=medium,top=thin)
       
    sh1.cell(row = 19,column=8).border=Border(left=thin,top=thin,bottom=thin)
    sh1.cell(row = 19,column=9).border=Border(right=thin,top=thin,bottom=thin,left=thin)
    sh1.cell(row = 19,column=7).border=Border(right=thin,top=thin,bottom=thin)
    sh1.cell(row = 19,column=2).border=Border(left=thin,top=thin,bottom=thin)
    for i in range(3,7):
        sh1.cell(row = 19,column=i).border=Border(top=thin,bottom=thin)
        
    sh1.cell(row = 19,column=1).border=Border(left=medium,top=thin,bottom=thin)
    
    sh1.cell(row = 19,column=10).border=Border(left=thin,top=thin,bottom=thin)
    sh1.cell(row = 19,column=11).border=Border(left=thin,right=medium,top=thin,bottom=thin)
    
    
    # sh1.cell(row = 19,column=12).border=Border(left=thin,top=thin,bottom=thin)
    # sh1.cell(row = 19,column=13).border=Border(top=thin,bottom=thin)
    # sh1.cell(row = 19,column=14).border=Border(right=medium,top=thin,bottom=thin)
    # ========================================================20 to 27 ==========================================
    sh1.cell(row = 25,column=1).border=Border(left=medium,top=thin,right=thin)
    sh1.cell(row = 26,column=1).border=Border(left=medium,right=thin)
    sh1.cell(row = 27,column=1).border=Border(left=medium,right=thin)
    sh1.cell(row = 28,column=1).border=Border(left=medium,bottom=thin,right=thin)
    
    
    # sh1.cell(row = 25,column=12).border=Border(left=thin,top=thin)
    # sh1.cell(row = 26,column=12).border=Border(left=thin)
    # sh1.cell(row = 27,column=12).border=Border(left=thin)
    # sh1.cell(row = 27,column=12).border=Border(left=thin,bottom=thin)
    
    
    sh1.cell(row = 25,column=11).border=Border(left=thin,right=medium,top=thin)
    sh1.cell(row = 26,column=11).border=Border(left=thin,right=medium)
    # sh1.cell(row = 27,column=14).border=Border(right=medium)
    sh1.cell(row = 27,column=11).border=Border(left=thin,right=medium,bottom=thin)
    
    


    for i in range(20,23):
        sh1.cell(row = i,column=1).border=Border(left=medium)
        sh1.cell(row = i,column=9).border=Border(left=thin)
        sh1.cell(row = i,column=10).border=Border(left=thin)
        # sh1.cell(row = i,column=12).border=Border(left=thin)
        sh1.cell(row = i,column = 11).border=Border(left=thin,right=medium)
    
    sh1.cell(row = 20,column=1).border=Border(right=thin,left=medium,top=thin)
    sh1.cell(row = 21,column=1).border=Border(right=thin,left=medium)
    sh1.cell(row = 22,column=1).border=Border(right=thin,left=medium)

    sh1.cell(row = 23,column=2).border=Border(left=thin,bottom=thin)
    for i in range(3,7):
        sh1.cell(row = 23,column=i).border=Border(bottom=thin)
    sh1.cell(row = 23,column=7).border=Border(bottom=thin)
    sh1.cell(row = 23,column=10).border=Border(right=thin,bottom=thin)
    sh1.cell(row = 23,column=11).border=Border(left=thin,right=medium,bottom=thin)
    # sh1.cell(row = 23,column=13).border=Border(bottom=thin)
       
    sh1.cell(row = 23,column=9).border=Border(left=thin,bottom=thin)
    sh1.cell(row = 23,column=10).border=Border(left=thin,bottom=thin)
    # sh1.cell(row = 23,column=12).border=Border(left=thin,bottom=thin)
    # sh1.cell(row = 23,column=11).border=Border(right=medium,bottom=thin)

    sh1.cell(row = 23,column=1).border=Border(left=medium,bottom=thin)
    
    sh1.cell(row = 24,column=1).border=Border(left=medium,bottom=thin,top=thin)
    sh1.cell(row = 24,column=2).border=Border(left=thin,bottom=thin,top=thin)
    for i in range(3,11):
        sh1.cell(row = 24,column=i).border=Border(top=thin,bottom=thin)
    # sh1.cell(row = 24,column=12).border=Border(left=thin,bottom=thin,top=thin)
    # sh1.cell(row = 24,column=13).border=Border(bottom=thin,top=thin)
    sh1.cell(row = 24,column=11).border=Border(left=thin,right=medium,bottom=thin,top=thin)
    
    # ============================================================= 29 =================================================

    sh1.cell(row = 28,column=11).border=Border(right=medium,bottom=thin,top=thin)
    for i in range(2,11):
        sh1.cell(row = 28,column=i).border=Border(bottom=thin,top=thin)
    sh1.cell(row = 28,column=11).border=Border(bottom=thin,top=thin,right=medium)
    
    sh1.cell(row = 28,column=1).border=Border(left=medium,top=thin,bottom=thin)

    # ======================================================= 30 =====================================================

    # sh1.cell(row = 29,column=15).border=Border(right=medium,bottom=thin,top=thin)
    
    # sh1.cell(row = 29,column=12).border=Border(left=thin,top=thin,bottom=thin)
    sh1.cell(row = 29,column=1).border=Border(left=medium,top=thin,bottom=thin)
    
    
    for i in range(2,11):
        sh1.cell(row = 29,column=i).border=Border(bottom=thin,top=thin)
    sh1.cell(row = 29,column=11).border=Border(bottom=thin,top=thin,right=medium)
    
    sh1.cell(row = 29,column=1).border=Border(left=medium,top=thin,bottom=thin)
    
    sh1.cell(row = 29,column=10).border=Border(left=thin,top=thin,bottom=thin)
    # =========================================================== 31 to 36 =======================================================

    # sh1.cell(row = 30,column=1).border=Border(left=medium,top=thin)

    # sh1.cell(row = 30,column=10).border=Border(left=thin,top=thin)

    # for i in range(30,34):
    #     sh1.cell(row = i,column=14).border=Border(right=medium)
    
    # sh1.cell(row = 34,column=14).border=Border(right=medium,bottom=medium)

    # for i in range(31,35):
    #     sh1.cell(row = i,column=1).border=Border(left=medium)
    #     sh1.cell(row = i,column=10).border=Border(left=thin)

    # for i in range(1,14):
    #     sh1.cell(row = 35,column=i).border=Border(top=medium)
    
# ====================================================Border Over=================================================            

    sh1.column_dimensions['I'].width = 17
    sh1.column_dimensions['J'].width = 17
    sh1.column_dimensions['H'].width = 17
    sh1.column_dimensions['G'].width = 17
    sh1.column_dimensions['E'].width = 10
    sh1.column_dimensions['F'].width = 17
    sh1.column_dimensions['D'].width = 14
    sh1.column_dimensions['K'].width = 19
    sh1.column_dimensions['A'].width = 10
    sh1.column_dimensions['B'].width = 15
    
    sh1.row_dimensions[i].ht = 40.0
    for i in range(2,30):
        sh1.row_dimensions[i].ht = 20.0

     
    
    # sh1.set_print_scale(50)
    openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(sh1, paper_size = 6, orientation='landscape')
    # sh1.page_setup.fitToWidth = 1
    #11
    # sh1.page_setup.fitToHeight = 0
    
    sh1.page_setup.fitToPage = True
    
    # sh1.print_area = 'A1:K29'

    # sh1.set_print_scale(90)
    sh1.page_margins.top=0
    sh1.page_margins.right=0.0
    sh1.page_margins.left=0.4
    sh1.page_margins.bottom=0.2
    sh1.page_margins.footer = 0.0
    sh1.page_margins.header = 0.0



    # os.chdir(u.BASEDIR)
    foldername = daten.strftime("%b_%y")
    
    dir = u.BASEDIR_BILL+"\\\\"+foldername
    try:
        if(os.path.isdir(dir) == False):
            os.mkdir(dir)
    except Exception as e:
        print("There is a error in creating folder : {0}".format(e))
        
    u.saveLocation = dir +r"\\"+str(u.bill_txt)+".xlsx"
    
    saveCustomerData(u,name=u.name_txt.get(),mobile=u.mobile_txt.get(),addhar_number=u.addhar_txt.get(),address=u.address_txt.get())
    # wb.save(filename='text.xlsx')
    # '''
    print(u.saveLocation)
    if(u.mobile_txt.get()!=''):
        saveBillLocation(u)
        print('save Location - ',u.saveLocation)
        try:
            wb.save(filename=u.saveLocation)
            print_bill(u)
        except Exception as e:
            messagebox.showerror("Error","Error in saving the Bill : {0}".format(e))
        
    
    
def print_bill(u:UiFields):
    try:
        # u.BASEDIR = findBASEDIR()
        # wb.save(filename=u.BASEDIR+'\test.xlsx')    
        thread = Thread(target = printBill,args=(u.saveLocation,))
        thread.start()
        # thread2 = Thread(target = printDialog)
        # thread2.start()
        # thread.join()
        print('Printing - ',"done printing")
        # thread2.join()
        # main_thread().sleep(100)
        time.sleep(4)
        subprocess.call(["taskkill","/F","/IM","excel.exe"])
    except Exception as e:
        print("Exception in printing : {0}".format(e))
        messagebox.showerror("Error","Error in printing the Bill : {0}".format(e))
    # '''
