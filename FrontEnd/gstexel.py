import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import os
import openpyxl
from openpyxl.styles import Border, Side, Font
from baseIntialization import UiFields

j = 4
wb = openpyxl.Workbook()
sh1 = wb.active
medium = Side(border_style='thin',color='FF000000')


def create(data, u:UiFields):
    
    sh1.title = 'gst'
    
    i = 0
    global j
    sh1['A2'] = 'id'
    sh1['B2'] = 'Date'
    sh1['C2'] = 'Ornament'
    sh1['D2'] = 'Qty'
    sh1['E2'] = 'Weight'
    sh1['F2'] = 'GoldRate'
    sh1['G2'] = 'Taxable Amt'
    sh1['H2'] = 'CGST'
    sh1['I2'] = 'SGST'
    sh1['J2'] = 'Net Total'
    for k in range(1,11):
        sh1.cell(row = 2,column = k).border=Border(left=medium,bottom=medium,top=medium)
        sh1.cell(row = 3,column = k).border=Border(bottom=medium)
    sh1.cell(row=1,column=10).border=Border(right=medium,left=medium,bottom=medium,top=medium)
    for gst in data:
        for c in gst:
            if(i==9):
                sh1.cell(row = j,column = i+1).border=Border(left=medium,bottom=medium,right=medium)
            else:
                sh1.cell(row = j,column = i+1).border=Border(left=medium,bottom=medium)
            location = chr(65+i)+''+str(j)
            if(i == 1):
                sh1[location] = str(c)[0:10]
            # print(location)
            else:
                sh1[location] = c
            i+=1
        j+=1
        i = 0
    sh1['C1'] = '{:%b %d, %Y}'.format(u.cal1)
    sh1['C1'].font = Font(name='times new rommon',size=18,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)
    sh1['E1'] = "TO"
    sh1['E1'].font = Font(name='times new rommon',size=18,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)
    sh1['G1'] = '{:%b %d, %Y}'.format(u.cal2)
    sh1['G1'].font = Font(name='times new rommon',size=18,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color=u.red)
    
    sh1.row_dimensions[1].ht = 40.0
    sh1.column_dimensions['A'].width = 15
    sh1.column_dimensions['B'].width = 15
    sh1.column_dimensions['C'].width = 15
    sh1.column_dimensions['D'].width = 15
    sh1.column_dimensions['E'].width = 15
    sh1.column_dimensions['F'].width = 15
    sh1.column_dimensions['G'].width = 15
    sh1.column_dimensions['H'].width = 15
    sh1.column_dimensions['I'].width = 15
    sh1.column_dimensions['J'].width = 15

    if os.path.exists(u.BASEDIR_GST)==False:
        os.mkdir(u.BASEDIR_GST)
        
    filename =os.path.join(u.BASEDIR_GST,"{:%b_%d_%Y}_{:%b_%d_%Y}.xlsx".format(u.cal1,u.cal2))
    
    wb.save(filename=filename)
    
def insertTotal(data, u:UiFields):
    global j
    k = j
    for k in range(j,j+3):
        for i in range(6,10):
            sh1.cell(row = k,column = i+1).border=Border(left=medium,bottom=medium,right=medium,top=medium)
    
    j = j+1
    
    sh1['G'+str(j)]=data[0][0]
    sh1['H'+str(j)]=data[0][1]
    sh1['I'+str(j)]=data[0][2]
    sh1['J'+str(j)]=data[0][3]
    
    j+=1
    
    sh1['G'+str(j)]='Taxable Amt'
    sh1['H'+str(j)]='CGST'
    sh1['I'+str(j)]='SGST'
    sh1['J'+str(j)]='NetTotal'
    
    if os.path.exists(u.BASEDIR_GST)==False:
        os.mkdir(u.BASEDIR_GST)
    filename =os.path.join(u.BASEDIR_GST,"{:%b_%d_%Y}_{:%b_%d_%Y}.xlsx".format(u.cal1,u.cal2))
    
    wb.save(filename=filename)
    send_email(filename, u)   
        
def send_email(attachment, u:UiFields):
    try:
        login = u.email_from_address
        password = u.email_from_pass
        smtpserver='smtp.gmail.com:587'
        from_addr = u.email_from_address
        to_addr_list = u.email_to_address
        print(to_addr_list)
        
        from_name = "GJ"
        plain_text_body = "GST Collection From {:%b %d, %Y} To {:%b %d, %Y}".format(u.cal1,u.cal2)
        subject = "GIRIDHARI JEWELLERY"
        
        
        message=MIMEMultipart()

        plain=MIMEText(plain_text_body,'plain')
        # html=MIMEText(html_body,'html') 

        message.add_header('from',from_name)
        message.add_header('to',to_addr_list)
        message.add_header('subject',subject)

        if attachment!=None:
            #attach_file=MIMEBase('application',"octet-stream")
            #attach_file.set_payload(open(attachment,"rb").read())
            #Encoders.encode_base64(attach_file)
            #f.close()
            attach_file=MIMEApplication(open(attachment,"rb").read())
            attach_file.add_header('Content-Disposition', 'attachment', filename=attachment)
            message.attach(attach_file)


        message.attach(plain)
        

        server = smtplib.SMTP(smtpserver)
        server.starttls()
        server.login(login,password)
        server.sendmail(from_addr, to_addr_list, message.as_string())
        server.quit()
    except Exception as e:
        print("There is a problem in sending the gst in email : {0}".format(e))
    