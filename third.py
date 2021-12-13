from tkinter import *
from  tkinter import ttk

from tkinter import font
from tkinter.ttk import *
from tkinter import messagebox
import tkinter
from tkinter import *
from tkinter import messagebox
import openpyxl
import tempfile
import os
import sys
from openpyxl.drawing.image import Image 
from openpyxl.styles import PatternFill,Border, Side, Alignment, Protection, Font, borders,fills
window = tkinter.Tk()
  
import datetime
daten = datetime.datetime.now()



# setting attribute
window.attributes('-fullscreen', True)
window.configure(bg='#FFE6BC')

window.title("Giridhari Jewellery")

def exit_(event):
    if(tkinter.messagebox.askokcancel('QUIT','Do You Want to Quit??')):
        window.withdraw()
        sys.exit() 
 

 

window.bind('<Escape>', exit_)

bg_color = "#FFE6BC"

# ==================================creating the gui================================================

Name=Label(window, text='Name and Address:', font=('times new rommon',10),bg=bg_color)
Name.grid(row=1,column=0,padx=10,pady=5)
Name_txt=Entry(window,width=30,font='arial 10',bd=2,justify=LEFT)
Name_txt.grid(row=1,column=1,pady=15)

mobile=Label(window, text='Mobile No.:', font=('times new rommon',10),bg=bg_color)
mobile.grid(row=1,column=4,padx=10,pady=5)
mobile_txt=Entry(window,width=20,font='arial 10',bd=2,justify=CENTER)
mobile_txt.grid(row=1,column=5,pady=15)

addhar=Label(window, text='Addhar No.:', font=('times new rommon',10),bg=bg_color)
addhar.grid(row=1,column=9,padx=20)
addhar_txt=Entry(window,width=25,font='arial 10',bd=2,justify=CENTER)
addhar_txt.grid(row=1,column=10)

bill=Label(window, text='Bill No.:', font=('times new rommon',10),bg=bg_color)
bill.grid(row=1,column=14,padx=10,pady=5)
bill_txt=Entry(window,width=15,font='arial 10',bd=2,justify=CENTER)
bill_txt.grid(row=1,column=15,pady=15)

date_label=Label(window, text=daten.strftime("%d-%b-%y - (%A)"), font=('times new rommon',10),bg=bg_color)
date_label.grid(row=1,column=20,padx=50,pady=5)


# ===========================================================================================================

F2 = LabelFrame(window,bg= "#FFE6BC")
F2.place(x=5, y=80,width=1355,height=370)

siLabel = Label(F2,text="Sino.",font=('times new rommon',10),bg=bg_color)
siLabel.grid(column=0,row=0)
si_txt = []
for i in range(1,10):
    txt=Label(F2,text=i,font=('times new rommon',10),bg=bg_color)
    txt.grid(row=i,column=0,padx=0,pady=5)
    si_txt.append(txt)


desLabel = Label(F2,text="Description",font=('times new rommon',10),bg=bg_color)
desLabel.grid(column=1,row=0)
des_txt = []

wtLabel = Label(F2,text="Weight",font=('times new rommon',10),bg=bg_color)
wtLabel.grid(column=2,row=0)
wt_txt = []

unitLabel = Label(F2,text="Unit Price",font=('times new rommon',10),bg=bg_color)
unitLabel.grid(column=3,row=0)
unitLabel_txt = []

cgstLabel = Label(F2,text="CGST(1.5%)",font=('times new rommon',10),bg=bg_color)
cgstLabel.grid(column=4,row=0)
cgstLabel_txt = []

sgstLabel = Label(F2,text="SGST(1.5%)",font=('times new rommon',10),bg=bg_color)
sgstLabel.grid(column=5,row=0)
sgstLabel_txt = []

toLabel = Label(F2,text="Net",font=('times new rommon',10),bg=bg_color)
toLabel.grid(column=6,row=0)
toLabel_txt = []

for i in range(1,10):
    txt1=Entry(F2,width=80,font='arial 12',bd=1,justify=CENTER)
    txt1.grid(row=i,column=1,padx=5,pady=5)
    des_txt.append(txt1)

    txt2=Entry(F2,width=9,font='arial 12',bd=1,justify=CENTER)
    txt2.grid(row=i,column=2,padx=5,pady=5)
    wt_txt.append(txt2)

    txt3=Entry(F2,width=9,font='arial 15',bd=1,justify=CENTER)
    txt3.grid(row=i,column=3,padx=5,pady=5)
    unitLabel_txt.append(txt3)

    txt4=Entry(F2,width=10,font='arial 12',bd=1,justify=CENTER)
    txt4.grid(row=i,column=4,padx=5,pady=5)
    cgstLabel_txt.append(txt4)

    txt5=Entry(F2,width=10,font='arial 12',bd=1,justify=CENTER)
    txt5.grid(row=i,column=5,padx=5,pady=5)
    sgstLabel_txt.append(txt5)

    txt6=Entry(F2,width=16,font='arial 12',bd=1,justify=CENTER)
    txt6.grid(row=i,column=6,padx=5,pady=5)
    toLabel_txt.append(txt6)




# ======================================================================================================================


F3 = LabelFrame(window,bg= "#FFE6BC")
F3.place(x=5,y=460,width=13055,height=100)

adrdeL = Label(F3,text="Old jewels",font=('times new rommon',10),bg=bg_color)
adrdeL.grid(column=0,row=0)
adrde_txt=Text(F3,width=100,height=3,font='arial 12')
adrde_txt.grid(row=1,column=0)

oldweLabel = Label(F3,text="Weight",font=('times new rommon',10),bg=bg_color)
oldweLabel.grid(column=1,row=0)
oldwe_txt=Entry(F3,width=15,font='arial 10',bd=1,justify=CENTER)
oldwe_txt.grid(row=1,column=1,padx=5,pady=5)

oldunitLabel = Label(F3,text="Unit Price",font=('times new rommon',10),bg=bg_color)
oldunitLabel.grid(column=2,row=0)
oldunit_txt=Entry(F3,width=15,font='arial 12',bd=1,justify=CENTER)
oldunit_txt.grid(row=1,column=2,padx=5,pady=5)


oldtotalLabel = Label(F3,text="Amount",font=('times new rommon',10),bg=bg_color)
oldtotalLabel.grid(column=3,row=0)
oldtotal_txt=Entry(F3,width=15,font='arial 12',bd=1,justify=CENTER)
oldtotal_txt.grid(row=1,column=3,padx=5,pady=5)

# =====================================================================================================================
F4 = LabelFrame(window,bg= "#FFE6BC")
F4.place(x=1100,y=570,width=250,height=80)

netotal_l = Label(F4,text="Total Paid",font=('times new rommon',12),bg=bg_color)
netotal_l.grid(column=3,row=0)
netotal= Entry(F4,width=15,font='arial 14',bd=1,justify=CENTER)
netotal.grid(row=0,column=4,padx=10,pady=5)



# ==========================================enter key binding===========================================

Name_txt.focus()

# define a function to change the tab order
# def tab_order(event):
#     widget = [Name_txt,addhar_txt,mobile_txt,bill_txt]
#     for w in widget:
#         w.lift()

# def enter(event):
#     print('Button-2 pressed at x = % d, y = % d'%(event.x, event.y))

# window.bind('<Return>', tab_order)





# ======================================function of the Code================================================
def generateBill():
    wb = openpyxl.Workbook()
    sh1 = wb.active

# ====================================================shop details==================================================================
    sh1['E1'] = "GIRIDHARI JEWELLERY"
    sh1['E1'].font = Font(name='times new rommon',size=25,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')

    sh1['D2'] = "Badheibanka Chawk, Old Town, Bhubaneswar, Mob.: 9090280083,9861230757"
    sh1['D2'].font = Font(name='times new rommon',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='064635')


    sh1['F3'] = "916 GOLD AND SILVER ORNAMENT"
    sh1['F3'].font = Font(name='times new rommon',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    
    sh1['G4'] = "Invoice"
    sh1['G4'].font = Font(name='times new rommon',size=14,bold=True,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')


    logo = Image("hallLogo.png")
    logo.height=100
    logo.width=100
    sh1.add_image(logo,'M1')

# ======================================================customer details==================================================

    green = '064635'
    red = 'C85C5C'
    blue = '041C32'
    # row_loc= row_loc+row_num+dis 
    sh1['A5'] = 'GSTIN - 21AYRPK4931F1ZH'
    sh1['A5'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')

    sh1['H5'] = "SI. No.: "
    sh1['H5'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='064635')
    sh1['I5'] = bill_txt.get()
    sh1['I5'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')

    sh1['K5'] = 'Date: '+ daten.strftime("%d-%b-%y - (%A)")
    sh1['K5'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='064635')

    # for i in range(1,col_num-1):
    #      sh1.cell(row=5, column=i+1).border=Border(top=Side(border_style='medium',color='FF000000'),
    #      bottom=Side(border_style='thin',color='FF000000'))
    
    # sh1.cell(row=5, column=1).border=Border(top=Side(border_style='medium',color='FF000000'),
    #      bottom=Side(border_style='thin',color='FF000000'),left=Side(border_style='medium',color='FF000000'))
    
    # sh1.cell(row=5, column=col_num).border=Border(top=Side(border_style='medium',color='FF000000'),
    #      bottom=Side(border_style='thin',color='FF000000'),right=Side(border_style='medium',color='FF000000'))

    sh1['A6'] = "Name: "
    sh1['A6'].font = Font(name='times new rommon',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['B6'] = Name_txt.get()
    sh1['B6'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)

    sh1['F6'] = 'Mobile No.'
    sh1['F6'].font = Font(name='times new rommon',size=11,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['G6'] = '  '+mobile_txt.get()
    sh1['G6'].font = Font(name='arial',size=10,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)

    sh1['I6'] = 'Addhar No.'
    sh1['I6'].font = Font(name='times new rommon',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['J6'] = addhar_txt.get()
    sh1['J6'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)

    sh1['M6'] = 'Bill No.'
    sh1['M6'].font = Font(name='times new rommon',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['N6'] = bill_txt.get()
    sh1['N6'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)

    sh1['F7'] = "Gold Rate: "
    sh1['F7'].font = Font(name='times new rommon',size=15,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['H7'] = "45000"
    sh1['H7'].font = Font(name='arial',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)


# ===========================================product details======================================================================
    
    sh1['A8'] = 'Si.no.'
    sh1['A8'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['B8'] = 'Description Of Goods'
    sh1['B8'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['F8'] = 'HSN'
    sh1['F8'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['F9'] = 'SAC'
    sh1['F9'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    
    sh1['G8'] = 'Weight'
    sh1['G8'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)

    sh1['H8'] = 'Unit'
    sh1['H8'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['H9'] = 'Price'
    sh1['H9'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)

    # sh1['J7'] = 'Taxable'
    # sh1['J7'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    # sh1['J7'] = 'Amount'
    # sh1['J7'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['I8'] = 'CGST'
    sh1['I8'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['I9'] = '1.5%'
    sh1['I9'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)

    sh1['K8'] = 'SGST'
    sh1['K8'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['K9'] = '1.5%'
    sh1['K9'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)

    sh1['M8'] = 'Net'
    sh1['M8'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)
    sh1['M9'] = 'Total'
    sh1['M9'].font = Font(name='times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)

    sh1['F10'] = '7113'
    sh1['F10'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

    for i in range(1,10):
        if(des_txt[i-1].get()!= ""):
            sh1['A'+str(i+9)] = i
            sh1['A'+str(i+9)].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

            sh1['B'+str(i+9)] = des_txt[i-1].get()
            sh1['B'+str(i+9)].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

            sh1['G'+str(i+9)] = wt_txt[i-1].get()
            sh1['G'+str(i+9)].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

            sh1['H'+str(i+9)] = unitLabel_txt[i-1].get()
            sh1['H'+str(i+9)].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

            # sh1['K'+str(i+8)] = [i-1].get()
            # sh1['N8'].font = Font(name='times new romman',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=green)

            sh1['I'+str(i+9)] = cgstLabel_txt[i-1].get()
            sh1['I'+str(i+9)].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

            sh1['K'+str(i+9)] = sgstLabel_txt[i-1].get()
            sh1['K'+str(i+9)].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

            sh1['M'+str(i+9)] = toLabel_txt[i-1].get()
            sh1['M'+str(i+9)].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)


# ===========================================================old gold============================================================

        sh1['D19'] = "Old Jewellery"
        # sh1['A20'] = adrde_txt.get()
        sh1['D19'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)

        sh1['I19'] = "Weight"
        sh1['I19'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)
        sh1['I21'] = oldwe_txt.get()
        sh1['I21'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)
        
        sh1['J19'] = "Unit Price"
        sh1['J19'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)
        sh1['J21'] = oldunit_txt.get()
        sh1['J21'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)


        sh1['M19'] = "Total"
        sh1['M19'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)
        sh1['M21'] = oldtotal_txt.get()
        sh1['M21'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

    # ===================================================mode of payment============================================


        sh1['E29'] = "Mode Of Payment"
        sh1['E29'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)

        sh1['I29'] = "Cash"
        sh1['I29'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

# =====================================================Total paid=======================================================

        sh1['A30'] = "Total in Words"
        sh1['A30'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)
        sh1['C30'] = "aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa"
        sh1['C30'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

        sh1['J30'] = "Total Paid : "
        sh1['J30'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)
        sh1['L30'] = '2000000'
        sh1['L30'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)
# =================================================Terms and conditions===================================================
        sh1['B31']="Terms and conditions"
        sh1['B31'].font = Font(name='Times new romman',size=14,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=red)

        sh1['A32'] = " 1. Before taking delivery Customer should check the ornaments by piece and weight. "
        sh1['A32'].font = Font(name='Times new romman',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

        sh1['A33'] = " 2. Exchange your ornaments within 7 days in good condition."
        sh1['A33'].font = Font(name='Times new romman',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

        sh1['A34'] = " 3. We are not responsible for any damage or brakage after delivery"
        sh1['A34'].font = Font(name='Times new romman',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)

        sh1['A35'] = " 4. We can repair the ornaments if possible"
        sh1['A35'].font = Font(name='Times new romman',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color=blue)


# ========================================================creating borders=====================================================
    # sh1['A3'] = "Name"
    thin_border_bottom = Border(bottom=Side(border_style='thin',color='FF000000'))
    thin_border_right = Border(right=Side(border_style='dashed',color='FF000000'))
    thin_border_top   = Border(top=Side(border_style='thin',color='FF000000'))
    thin_border_left  = Border(left=Side(border_style='thin',color='FF000000'))
                
    thick_border_left = Border(left=Side(border_style='medium',color='FF000000'))
    thick_border_right = Border(right=Side(border_style='medium',color='FF000000'))
    thick_border_top = Border(top=Side(border_style='medium',color='FF000000'))
    thick_border_bottom = Border(bottom=Side(border_style='medium',color='FF000000'))
                    

    fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00',end_color='00FFFF00')

    # ================================================== 5 =====================================================
    for i in range(2,14):
        sh1.cell(row=5, column=i).border=Border(top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row=5,column=1).border=Border(left=Side(border_style='medium',color='FF000000'),top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row=5,column=14).border=Border(right=Side(border_style='medium',color='FF000000'),top=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    
    # ==================================================== 6 =====================================================

    sh1.cell(row = 6,column = 1).border=Border(left=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 6,column = 14).border=Border(right=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'))

    # ======================================================= 7 ====================================================

    sh1.cell(row = 7,column = 14).border=Border(right=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 7,column = 1).border=Border(left=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))


    # ======================================================= 8 =====================================================

    sh1.cell(row = 8,column = 1).border=Border(left=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 2).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 3).border=Border(top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 4).border=Border(top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 5).border=Border(right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))

    sh1.cell(row = 8,column = 6).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 7).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 8).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 9).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 10).border=Border(right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))

    sh1.cell(row = 8,column = 11).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 12).border=Border(right=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 13).border=Border(top=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 8,column = 14).border=Border(right=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    
    # ===================================================== 9 ======================================================
    
    sh1.cell(row = 9,column = 1).border=Border(left=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 2).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 3).border=Border(bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 4).border=Border(bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 5).border=Border(bottom=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 6).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 7).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 8).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 9).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 10).border=Border(right=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 11).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),)
    sh1.cell(row = 9,column = 12).border=Border(right=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 13).border=Border(bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 9,column = 14).border=Border(right=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))


    # ========================================================== 10 =======================================================

    sh1.cell(row = 10,column = 1).border = Border(left=Side(border_style='medium',color='FF000000'))
    for i in range(2,15):
        if(i!=3 and i!=4 and i!=5 and i!=10 and i!=12):
            sh1.cell(row = 10,column = i).border = Border(left=Side(border_style='thin',color='FF000000'))

    # ============================================================= 10 to 18 ===================================================
    
    for i in range(10,18):
        sh1.cell(row = i,column = 14).border=Border(right=Side(border_style='medium',color='FF000000'))

    for i in range(11,18):
        sh1.cell(row = i,column = 1).border = Border(left=Side(border_style='medium',color='FF000000'))

    for j in range(1,15):
        if(j==2 or j==6 or j==7 or j==8 or j==9 or j==11 or j==13):  
            for i in range(11,18):
                sh1.cell(row = i,column = j).border = Border(left=Side(border_style='thin',color='FF000000'))

    sh1.cell(row = 18,column = 1).border=Border(left=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'))

    sh1.cell(row = 18,column = 2).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 3).border=Border(bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 4).border=Border(bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 5).border=Border(bottom=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 6).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 7).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 8).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 9).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 10).border=Border(bottom=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 11).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 12).border=Border(bottom=Side(border_style='thin',color='FF000000'),right=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 18,column = 13).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))    
    sh1.cell(row = 18,column = 14).border=Border(right=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))    
    

    # ======================================================= 19 ==========================================================

    sh1.cell(row = 19,column = 14).border=Border(right=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'))    
    sh1.cell(row = 19,column=8).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))

    sh1.cell(row = 19,column=1).border=Border(left=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    
    sh1.cell(row = 19,column=10).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    
    sh1.cell(row = 19,column=12).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    # ========================================================20 to 27 ==========================================
    sh1.cell(row = 26,column=1).border=Border(left=Side(border_style='medium',color='FF000000'))


    for i in range(20,28):
        sh1.cell(row = i,column=1).border=Border(left=Side(border_style='medium',color='FF000000'))
        sh1.cell(row = i,column=8).border=Border(left=Side(border_style='thin',color='FF000000'))
        sh1.cell(row = i,column=10).border=Border(left=Side(border_style='thin',color='FF000000'))
        sh1.cell(row = i,column=12).border=Border(left=Side(border_style='thin',color='FF000000'))
        sh1.cell(row = i,column = 14).border=Border(right=Side(border_style='medium',color='FF000000'))

    
    sh1.cell(row = 28,column=8).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 28,column=10).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 28,column=12).border=Border(left=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 28,column=14).border=Border(right=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

    sh1.cell(row = 28,column=1).border=Border(left=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    
    # ============================================================= 29 =================================================

    sh1.cell(row = 29,column=14).border=Border(right=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    for i in range(2,14):
        sh1.cell(row = 29,column=i).border=Border(bottom=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    
    sh1.cell(row = 29,column=1).border=Border(left=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))

    # ======================================================= 30 =====================================================

    sh1.cell(row = 30,column=14).border=Border(right=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    
    sh1.cell(row = 30,column=12).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    sh1.cell(row = 30,column=1).border=Border(left=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    
    
    for i in range(2,14):
        sh1.cell(row = 30,column=i).border=Border(bottom=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))
    
    sh1.cell(row = 30,column=1).border=Border(left=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    
    sh1.cell(row = 30,column=10).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'),bottom=Side(border_style='thin',color='FF000000'))
    # =========================================================== 31 to 36 =======================================================

    sh1.cell(row = 31,column=1).border=Border(left=Side(border_style='medium',color='FF000000'),top=Side(border_style='thin',color='FF000000'))

    sh1.cell(row = 31,column=10).border=Border(left=Side(border_style='thin',color='FF000000'),top=Side(border_style='thin',color='FF000000'))

    for i in range(31,35):
        sh1.cell(row = i,column=14).border=Border(right=Side(border_style='medium',color='FF000000'))
    
    sh1.cell(row = 35,column=14).border=Border(right=Side(border_style='medium',color='FF000000'),bottom=Side(border_style='medium',color='FF000000'))

    for i in range(32,36):
        sh1.cell(row = i,column=1).border=Border(left=Side(border_style='medium',color='FF000000'))
        sh1.cell(row = i,column=10).border=Border(left=Side(border_style='thin',color='FF000000'))

    for i in range(1,14):
        sh1.cell(row = 36,column=i).border=Border(top=Side(border_style='medium',color='FF000000'))
    
# ====================================================Border Over=================================================            

            
    

    openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(sh1, paper_size = 1, orientation='landscape')
    sh1.page_margins.top=.2
    sh1.page_margins.right=.2
    sh1.page_margins.bottom=.2



    wb.save(filename='test.xlsx')
    os.system('test.xlsx')


def prin(event):
    os.startfile('test.xlsx','print')

def opena():
    print("Hello")
    os.system('test.xlsx')
# =================================================================================================================

F5 = LabelFrame(window,bg= "#519259")
F5.place(x=5,y=570,width=600,height=120)
print(des_txt[0].get() == "")

newBtn = Button(F5,text="New (Ctrl+N)",font=('times new rommon',13),bg=bg_color,bd=2)
newBtn.grid(column=0,row=0,padx=20,pady=10)

printBtn = Button(F5,text="Print (Ctrl+P)",font=('times new rommon',13),command=prin,bg=bg_color,bd=2)
printBtn.grid(column=0,row=1,padx=20,pady=10)

generateBtn = Button(F5,text="Generate Bill (Ctrl+G)",font=('times new rommon',13),command=generateBill,bg=bg_color,bd=2)
generateBtn.grid(column=1,row=0,padx=20,pady=10)

findBtn = Button(F5,text = "Find (Ctrl+F)",font=('times new rommon',13),command=open,bg=bg_color,bd=2)
findBtn.grid(column=1,row=1,padx=20,pady=10)



window.bind('<Control-G>', generateBill)
window.bind('<Control-p>', prin)
window.bind('<Control-slash>', opena)

# ========================================end of the code=========================================================================
window.mainloop()