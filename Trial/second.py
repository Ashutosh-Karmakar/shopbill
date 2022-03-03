from tkinter import *
from  tkinter import ttk

from tkinter import font
from tkinter.ttk import *
from tkinter import messagebox
import tkinter
from tkinter import *
from tkinter import messagebox
import openpyxl
import tempfile
import os
from openpyxl.styles import PatternFill,Border, Side, Alignment, Protection, Font,fills


window = tkinter.Tk()
#title of the window
window.title("GIRIDHARI JEWELLERY")
#size of the window
window.geometry('1920x1080')
# window.configure(bg='#FEFBF3')

window.configure(bg='#FEFBF3')


bg_color = "#FFE3E3"

# title=Label(window,text="GIRIDHARI JEWELLERY",bd=12,bg=bg_color,fg='#E63E6D',font=('times new roman', 30 ,'bold'),relief=GROOVE,justify=CENTER)
# title.pack(fill=X)

# #===============FRAME ONE=================
F1 = LabelFrame(window,bg= "#FFE3E3")
F1.place(x=5, y=5,width=1500,height=100)


#=====================PERTIONAL DETAILS==========================

Name=Label(F1, text='Name and Address:', font=('times new rommon',15),bg=bg_color)
Name.grid(row=1,column=0,padx=10,pady=5)
Name_txt=Entry(F1,width=30,font='arial 15',bd=2,justify=CENTER)
Name_txt.grid(row=1,column=1,pady=15)

date=Label(F1, text='Date:', font=('times new rommon',15),bg=bg_color)
date.grid(row=1,column=2,padx=10,pady=5)
date_txt=Entry(F1,width=10,font='arial 15',bd=2,justify=CENTER)
date_txt.grid(row=1,column=3,pady=15)

mobile=Label(F1, text='Mobile No.:', font=('times new rommon',15),bg=bg_color)
mobile.grid(row=1,column=4,padx=10,pady=5)
mobile_txt=Entry(F1,width=20,font='arial 15',bd=2,justify=CENTER)
mobile_txt.grid(row=1,column=5,pady=15)

bill=Label(F1, text='Bill No.:', font=('times new rommon',15),bg=bg_color)
bill.grid(row=1,column=6,padx=10,pady=5)
bill_txt=Entry(F1,width=15,font='arial 15',bd=2,justify=CENTER)
bill_txt.grid(row=1,column=7,pady=15)


pan=Label(F1, text='Pan No.:', font=('times new rommon',15),bg=bg_color)
pan.grid(row=3,column=0,padx=20)
pan_txt=Entry(F1,font='arial 15',bd=2,justify=CENTER)
pan_txt.grid(row=3,column=1)


gst=Label(F1, text='GST No.:', font=('times new rommon',15),bg=bg_color)
gst.grid(row=3,column=2,padx=20)
gst_txt=Entry(F1,width=10,font='arial 15',bd=2,justify=CENTER)
gst_txt.grid(row=3,column=3)

addhar=Label(F1, text='Addhar No.:', font=('times new rommon',15),bg=bg_color)
addhar.grid(row=3,column=4,padx=20)
addhar_txt=Entry(F1,width=25,font='arial 15',bd=2,justify=CENTER)
addhar_txt.grid(row=3,column=5)

gstin=Label(F1, text='GSTIN:', font=('times new rommon',15),bg=bg_color)
gstin.grid(row=3,column=6)
gstin=Label(F1, text='21AYRPK4931F1ZH', font=('times new rommon',15),bg=bg_color)
gstin.grid(row=3,column=7)

#===============FRAME ONE=================
F2 = LabelFrame(window,bg= "#FFE3E3")
F2.place(x=5, y=100,width=1500,height=425)

siLabel = Label(F2,text="Sino.",font=('times new rommon',15),bg=bg_color)
siLabel.grid(column=0,row=0)
si_txt = []
for i in range(1,11):
    txt=Entry(F2,width=5,font='arial 15',bd=2,justify=CENTER)
    txt.grid(row=i,column=0,padx=5,pady=5)
    si_txt.append(txt)

desLabel = Label(F2,text="Description",font=('times new rommon',15),bg=bg_color)
desLabel.grid(column=1,row=0)
des_txt = []
for i in range(1,11):
    txt=Entry(F2,width=40,font='arial 15',bd=2,justify=CENTER)
    txt.grid(row=i,column=1,padx=5,pady=5)
    des_txt.append(txt)

wtLabel = Label(F2,text="Weight",font=('times new rommon',15),bg=bg_color)
wtLabel.grid(column=2,row=0)
wt_txt = []
for i in range(1,11):
    txt=Entry(F2,width=10,font='arial 15',bd=2,justify=CENTER)
    txt.grid(row=i,column=2,padx=5,pady=5)
    wt_txt.append(txt)

unitLabel = Label(F2,text="Unit Price",font=('times new rommon',15),bg=bg_color)
unitLabel.grid(column=3,row=0)
unitLabel_txt = []
for i in range(1,11):
    txt=Entry(F2,width=10,font='arial 15',bd=2,justify=CENTER)
    txt.grid(row=i,column=3,padx=5,pady=5)
    unitLabel_txt.append(txt)

cgstLabel = Label(F2,text="CGST(1.5%)",font=('times new rommon',15),bg=bg_color)
cgstLabel.grid(column=4,row=0)
cgstLabel_txt = []
for i in range(1,11):
    txt=Entry(F2,width=15,font='arial 15',bd=2,justify=CENTER)
    txt.grid(row=i,column=4,padx=5,pady=5)
    cgstLabel_txt.append(txt)

sgstLabel = Label(F2,text="SGST(1.5%)",font=('times new rommon',15),bg=bg_color)
sgstLabel.grid(column=5,row=0)
sgstLabel_txt = []
for i in range(1,11):
    txt=Entry(F2,width=15,font='arial 15',bd=2,justify=CENTER)
    txt.grid(row=i,column=5,padx=5,pady=5)
    sgstLabel_txt.append(txt)

toLabel = Label(F2,text="Net",font=('times new rommon',15),bg=bg_color)
toLabel.grid(column=6,row=0)
toLabel_txt = []
for i in range(1,11):
    txt=Entry(F2,width=20,font='arial 15',bd=2,justify=CENTER)
    txt.grid(row=i,column=6,padx=5,pady=5)
    toLabel_txt.append(txt)
# ==================================================================


F3 = LabelFrame(window,bg= "#FFE3E3")
F3.place(x=5,y=525,width=1500,height=100)
adrdeL = Label(F3,text="Addition Or Deduction",font=('times new rommon',10),bg=bg_color)
adrdeL.grid(column=0,row=0)
adrde_txt=Text(F3,width=110,height=3,font='arial 15')
adrde_txt.grid(row=1,column=0)


amtL = Label(F3,text="Amount",font=('times new rommon',10),bg=bg_color)
amtL.grid(column=1,row=0)
amt_txt=Text(F3,width=20,height=3,font='arial 15',bd=2)
amt_txt.grid(row=1,column=1)
    

# ===============================================================
F4 = LabelFrame(window,bg= "#FFE3E3")
F4.place(x=5,y=625,width=1500,height=100)

modeL = Label(F4,text="Mode Of Payment:",font=('times new rommon',15),bg=bg_color)
modeL.grid(column=0,row=0)
mode_txt=Entry(F4,width=60,font='arial 15',bd=2)
mode_txt.grid(row=0,column=1,padx=10)

disL = Label(F4,text="Discount:",font=('times new rommon',15),bg=bg_color)
disL.grid(column=2,row=0,padx=10)
dis_txt=Entry(F4,width=20,font='arial 15',bd=2)
dis_txt.grid(row=0,column=3)

tamtL = Label(F4,text="Total Amount Paid:",font=('times new rommon',15),bg=bg_color)
tamtL.grid(column=2,row=1,padx=10)
tamt_txt=Entry(F4,width=20,font='arial 15',bd=2)
tamt_txt.grid(row=1,column=3,pady=10)




#============================================================================================
def generateBill():
    # billArea = tkinter.Tk()
    # billArea.title("Bill Area")
    # billArea.geometry('1920x1080')
    # bill = Text(billArea,width='1920',height='1080')
    # bill.grid(column=0,row=0)
    # # bill.configure(font='helvetica 15')
    # bill.insert("1.0","\t\t\t\t\t\tGIRIDHARI JEWELLERY")
    # bill.tag_add("Heading", "1.0", "1.50")  
    # # bill.tag_add("Click Here", "1.8", "1.13")  
  
    # bill.tag_config("Heading", foreground="#E63E6D",font=('times new rommon', 35))
    
    # bill.insert(END,"\nName & Address: " + Name_txt.get())
    # bill.tag_add("name","2.0","2.15")
    # bill.tag_config("name", foreground="#E63E6D",font=('times new rommon', 20))
    
    # bill.insert("2.50","\t\t\t\tDate: "+date_txt.get())
    # bill.tag_add("name","2.50","2.54")
    # bill.tag_config("name", foreground="#E63E6D",font=('times new rommon', 20))
    # bill.insert(END,"\t\t\t\tMobile No.: "+mobile_txt.get())
    # bill.insert(END,"\tBill No.: "+bill_txt.get())
    
    
    
    # billArea.mainloop()
    
    wb = openpyxl.Workbook()
    sh1 = wb.active
    sh1['D1'] = "GIRIDHARI JEWELLERY"
    font = Font(name='times new rommon',size=25,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    sh1['D1'].font = font
    
    sh1['A3'] = "NAME:"
    sh1['A3'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    sh1['B3'] = Name_txt.get()
    sh1['B3'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='2C272E')
    sh1['E3'] = "Date:"
    sh1['E3'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    sh1['F3'] = date_txt.get()
    sh1['F3'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='2C272E')
    sh1['H3'] = "Mobile No.:"
    sh1['H3'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    sh1['J3'] = mobile_txt.get()
    sh1['J3'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='2C272E')
    sh1['L3'] = "Bill No.:"
    sh1['L3'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    sh1['M3'] = bill_txt.get()
    sh1['M3'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='2C272E')
    
    sh1['A4'] = "Pan No.:"
    sh1['A4'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    sh1['B4'] = pan_txt.get()
    sh1['B4'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='2C272E')
    sh1['E4'] = "Gst No.:"
    sh1['E4'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    sh1['F4'] = gst_txt.get()
    sh1['F4'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='2C272E')
    sh1['H4'] = "Addhar No.:"
    sh1['H4'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    sh1['J4'] = addhar_txt.get()
    sh1['J4'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='2C272E')
    sh1['L4'] = "GSTIN:"
    sh1['L4'].font = Font(name='times new rommon',size=13,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='C85C5C')
    sh1['M4'] = "21AYRPK4931F1ZH"
    sh1['M4'].font = Font(name='arial',size=12,bold=False,italic=False,vertAlign=None,underline='none',strike=False,color='2C272E')
    
    thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
                right=Side(border_style='dashed',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000')
                )
                
    thick_border = Border(left=Side(border_style='thin',color='FF000000'),
                    right=Side(border_style='thin',color='FF000000'),
                    top=Side(border_style='thin',color='FF000000'),
                    bottom=Side(border_style='medium',color='FF000000')
                    )
    Double_border = Border(left=Side(border_style='dashed',color='FF000000'),
                    right=Side(border_style='dashed',color='FF000000'),
                    top=Side(border_style='double',color='FF000000'),
                    bottom=Side(border_style='double',color='FF000000')
                    )
    fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00',end_color='00FFFF00')
    #define size of the table 
    row_num=10
    col_num=7 
    #location of the Table 
    row_loc=6
    col_loc=6
    dis=2 # distance between the tables 
    for i in range (row_loc,row_loc+row_num):
        for j in range (col_loc,col_num+col_loc):
            sh1.cell(row=i+1, column=j+2).border=thin_border
            if i==row_loc:
                sh1.cell(row=i+1, column=j+2).border=Double_border
                sh1.cell(row=i+1, column=j+2).fill=fill_cell
            if i==row_loc+row_num-1:
                sh1.cell(row=i+1, column=j+2).border=thick_border
    row_loc= row_loc+row_num+dis 

    
    openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(sh1, paper_size = 1, orientation='landscape')
    sh1.page_margins.top=.2


    
    wb.save(filename='test.xlsx')
def prin():
    os.startfile('test.xlsx','print')
    
    
    

# ==============================================================
F5 = LabelFrame(window,bg= "#FFE3E3")
F5.place(x=5, y=730,width=1500,height=50)


newBtn = Button(F5,text="New",font=('times new rommon',10),bg=bg_color,bd=5)
newBtn.grid(column=0,row=0,padx=30)

printBtn = Button(F5,text="Print",command=prin,font=('times new rommon',10),bg=bg_color,bd=5)
printBtn.grid(column=1,row=0,padx=30)

generateBtn = Button(F5,text="Generate Bill",font=('times new rommon',10),command = generateBill,bg=bg_color,bd=5)
generateBtn.grid(column=2,row=0,padx=30)

findBtn = Button(F5,text = "Find",font=('times new rommon',10),bg=bg_color,bd=5)
findBtn.grid(column=3,row=0,padx=30)













window.mainloop()